import pandas as pd
import openpyxl as pyxl
from openpyxl import styles
    
def set_column_width(excelFilePath):
    # open Excel file
    workBook = pyxl.load_workbook(excelFilePath)
    workSheetCount = len(workBook.sheetnames)

    # set column width
    for i in range(workSheetCount):
        workSheet = workBook.worksheets[i]
        for col in workSheet.columns:
            max_length = 0

            if type(col[0]).__name__ == 'MergedCell':
                continue
            else:
                column = col[0].column_letter

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            workSheet.column_dimensions[column].width = adjusted_width

    # save Excel file
    workBook.save(excelFilePath)


def set_title_style(columns):
    for col in columns:
        for cell in col:
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')
            cell.font = styles.Font(bold=True)


def set_border(columns):
    for col in columns:
        for cell in col:
            cell.border = styles.Border(left=styles.Side(border_style='thin', color='000000'),
                                        right=styles.Side(border_style='thin', color='000000'),
                                        top=styles.Side(border_style='thin', color='000000'),
                                        bottom=styles.Side(border_style='thin', color='000000'))


def generate_excel(excelFilePath, memoryMapList, groupInfoList, subGroupInfoList):
    # make groupInfoList to dataFrame
    groupDf = pd.DataFrame(groupInfoList)

    # make subGroupInfoList to dataFrame
    subGroupDf = pd.DataFrame(subGroupInfoList)

    # make memoryMapList to dataFrame
    memoryMapDf = pd.DataFrame(index=[0], columns=['subGroupName', 'objFileName', 'objSize', 'address'])
    for memMap in memoryMapList:
        df = pd.DataFrame({'section' : memMap.section, 'groupName' : memMap.swComponent.group, 'subGroupName' : memMap.swComponent.subGroup,
                            'objFileName' : memMap.fileName, 'objSize' : memMap.objSize, 'address' : memMap.address},
                            index=[0])                
        memoryMapDf = pd.concat([memoryMapDf, df])

    # merge dataFrames, order by group
    mergedDf = pd.merge(groupDf, subGroupDf, how='outer', on='groupName')
    mergedDf = pd.merge(mergedDf, memoryMapDf, how='outer', on=['groupName', 'subGroupName'])

    # make dataFrame to Excel file
    mergedDf.to_excel(excelFilePath, index=False, sheet_name='All Datas')

    # set column width on Excel file
    set_column_width(excelFilePath)


def generate_group_info_sheet(excelFilePath, groupInfoList):
    # load excel file
    workBook = pyxl.load_workbook(excelFilePath)

    # create new sheet
    workBook.create_sheet(title='Group Memory Info')

    # set workSheet
    workSheet = workBook['Group Memory Info']

    # add title
    workSheet.cell(1,1).value = 'groupName'
    workSheet.cell(1,2).value = 'ROM'
    workSheet.cell(1,5).value = 'RAM'

    columns = ['groupName', '.text', '.rodata', 'romSize', '.data', '.bss', 'ramSize', 'groupSize']
    workSheet.append(columns)

    # merge title cells
    workSheet.merge_cells('B1:D1') # ROM
    workSheet.merge_cells('E1:G1') # RAM
    workSheet.merge_cells('A1:A2') # groupName

    # set title column style
    set_title_style(workSheet.columns)
    
    # calculate RAM/ROM size and write to workSheet
    totalRamSize = 0
    totalRomSize = 0
    totalSize = 0

    for gInfo in groupInfoList:
        romSize = gInfo['.text'] + gInfo['.rodata']
        ramSize = gInfo['.data'] + gInfo['.bss']
        groupSize = romSize + ramSize

        totalRamSize += ramSize
        totalRomSize += romSize
        totalSize += (romSize + ramSize)

        workSheet.append([gInfo['groupName'], gInfo['.text'], gInfo['.rodata'], romSize, gInfo['.data'], gInfo['.bss'], ramSize, groupSize])              

    # set cell border
    set_border(workSheet.columns)

    # append total RAM/ROM/total size
    workSheet.append([])
    workSheet.append(['', '', '', totalRomSize, '', '', totalRamSize, totalSize])

    # save workSheet
    workBook.save(excelFilePath)

    # set column width on Excel file
    set_column_width(excelFilePath)


def generate_subGroup_info_sheet(excelFilePath, subGroupInfoList):
    # load excel file
    workBook = pyxl.load_workbook(excelFilePath)

    # create new sheet
    workBook.create_sheet(title='Module Memory Info')

    # set workSheet
    workSheet = workBook['Module Memory Info']

    # add title
    workSheet.cell(1,1).value = 'moduleName'
    workSheet.cell(1,2).value = 'ROM'
    workSheet.cell(1,5).value = 'RAM'

    columns = ['moduleName', '.text', '.rodata', 'romSize', '.data', '.bss', 'ramSize', 'moduleSize']
    workSheet.append(columns)

    # merge title cells
    workSheet.merge_cells('B1:D1') # ROM
    workSheet.merge_cells('E1:G1') # RAM
    workSheet.merge_cells('A1:A2') # moduleName

    # set title column style
    set_title_style(workSheet.columns)

    # calculate RAM/ROM size and write to workSheet
    totalRamSize = 0
    totalRomSize = 0

    for sInfo in subGroupInfoList:
        romSize = sInfo['.text'] + sInfo['.rodata']
        ramSize = sInfo['.data'] + sInfo['.bss']
        subGroupSize = romSize + ramSize

        totalRomSize += romSize 
        totalRamSize += ramSize 

        workSheet.append([sInfo['subGroupName'], sInfo['.text'], sInfo['.rodata'], romSize, sInfo['.data'], sInfo['.bss'], ramSize, subGroupSize])              

    # set cell border
    set_border(workSheet.columns)

    # append total RAM/ROM size
    workSheet.append([])
    workSheet.append(['', '', '', totalRomSize, '', '', totalRamSize])

    # append notice
    workSheet.append([])
    workSheet.append(['', '', '', '', '', '', '', '', '', '* Since the etc group does not have modules (subGroup), the total RAM/ROM size does not include the size of the etc group.'])
    
    # save workSheet
    workBook.save(excelFilePath)

    # set column width on Excel file
    set_column_width(excelFilePath)